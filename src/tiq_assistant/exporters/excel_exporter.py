"""Excel export functionality for timesheet entries."""

from datetime import datetime
from pathlib import Path
from typing import Optional
import xlsxwriter
import openpyxl

from tiq_assistant.core.models import TimesheetEntry
from tiq_assistant.core.exceptions import ExportError
from tiq_assistant.storage.sqlite_store import get_store


class ExcelExporter:
    """Export timesheet entries to Excel format."""

    # Column configuration matching the exact timesheet format
    # Note: Column names have trailing spaces to match the template exactly
    COLUMNS = [
        ("Consultant ID", 20),
        ("Date         ", 15),
        ("Workhour     ", 10),
        ("Ticket No    ", 15),
        ("Project      ", 35),
        ("Activity No  ", 12),
        ("Location     ", 12),
        ("             ", 5),  # Empty spacer column (13 spaces)
        ("Activity", 50),
    ]

    # Mapping from internal keys to column headers (for to_export_row compatibility)
    COLUMN_KEY_MAP = {
        "Consultant ID": "Consultant ID",
        "Date": "Date         ",
        "Workhour": "Workhour     ",
        "Ticket No": "Ticket No    ",
        "Project": "Project      ",
        "Activity No": "Activity No  ",
        "Location": "Location     ",
        "Activity": "Activity",
    }

    def export_to_new_file(
        self,
        entries: list[TimesheetEntry],
        output_path: Path | str,
    ) -> Path:
        """Create a new Excel file with timesheet entries."""
        output_path = Path(output_path)

        try:
            workbook = xlsxwriter.Workbook(str(output_path))
            worksheet = workbook.add_worksheet("Timesheet")

            # Create formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
            })

            cell_format = workbook.add_format({
                'border': 1,
                'align': 'left',
                'valign': 'vcenter',
            })

            number_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
            })

            # Write header row
            for col, (name, width) in enumerate(self.COLUMNS):
                worksheet.write(0, col, name, header_format)
                worksheet.set_column(col, col, width)

            # Write data rows
            for row, entry in enumerate(entries, start=1):
                export_row = entry.to_export_row()
                worksheet.write(row, 0, export_row["Consultant ID"], cell_format)
                worksheet.write(row, 1, export_row["Date"], cell_format)
                worksheet.write(row, 2, export_row["Workhour"], number_format)
                worksheet.write(row, 3, export_row["Ticket No"], cell_format)
                worksheet.write(row, 4, export_row["Project"], cell_format)
                worksheet.write(row, 5, export_row["Activity No"], cell_format)
                worksheet.write(row, 6, export_row["Location"], cell_format)
                worksheet.write(row, 7, "", cell_format)  # Empty column
                worksheet.write(row, 8, export_row["Activity"], cell_format)

            # Freeze header row
            worksheet.freeze_panes(1, 0)

            workbook.close()
            return output_path

        except Exception as e:
            raise ExportError(f"Failed to create Excel file: {e}")

    def export_to_template(
        self,
        entries: list[TimesheetEntry],
        template_path: Path | str,
        output_path: Path | str,
        start_row: int = 2,  # Row 1 is header (1-indexed in openpyxl)
    ) -> Path:
        """Export entries to an existing template file."""
        template_path = Path(template_path)
        output_path = Path(output_path)

        if not template_path.exists():
            raise ExportError(f"Template file not found: {template_path}")

        try:
            # Load template
            workbook = openpyxl.load_workbook(template_path)
            worksheet = workbook.active

            # Write data starting from start_row
            for idx, entry in enumerate(entries):
                row = start_row + idx
                export_row = entry.to_export_row()

                worksheet.cell(row=row, column=1, value=export_row["Consultant ID"])
                worksheet.cell(row=row, column=2, value=export_row["Date"])
                worksheet.cell(row=row, column=3, value=export_row["Workhour"])
                worksheet.cell(row=row, column=4, value=export_row["Ticket No"])
                worksheet.cell(row=row, column=5, value=export_row["Project"])
                worksheet.cell(row=row, column=6, value=export_row["Activity No"])
                worksheet.cell(row=row, column=7, value=export_row["Location"])
                worksheet.cell(row=row, column=8, value="")  # Empty column
                worksheet.cell(row=row, column=9, value=export_row["Activity"])

            workbook.save(output_path)
            return output_path

        except Exception as e:
            raise ExportError(f"Failed to export to template: {e}")

    def append_to_existing(
        self,
        entries: list[TimesheetEntry],
        existing_path: Path | str,
    ) -> Path:
        """Append entries to an existing timesheet file."""
        existing_path = Path(existing_path)

        if not existing_path.exists():
            raise ExportError(f"File not found: {existing_path}")

        try:
            workbook = openpyxl.load_workbook(existing_path)
            worksheet = workbook.active

            # Find first empty row
            start_row = worksheet.max_row + 1

            # Write data
            for idx, entry in enumerate(entries):
                row = start_row + idx
                export_row = entry.to_export_row()

                worksheet.cell(row=row, column=1, value=export_row["Consultant ID"])
                worksheet.cell(row=row, column=2, value=export_row["Date"])
                worksheet.cell(row=row, column=3, value=export_row["Workhour"])
                worksheet.cell(row=row, column=4, value=export_row["Ticket No"])
                worksheet.cell(row=row, column=5, value=export_row["Project"])
                worksheet.cell(row=row, column=6, value=export_row["Activity No"])
                worksheet.cell(row=row, column=7, value=export_row["Location"])
                worksheet.cell(row=row, column=8, value="")  # Empty column
                worksheet.cell(row=row, column=9, value=export_row["Activity"])

            workbook.save(existing_path)
            return existing_path

        except Exception as e:
            raise ExportError(f"Failed to append to file: {e}")

    def entries_to_dataframe(self, entries: list[TimesheetEntry]):
        """Convert entries to a pandas DataFrame for display."""
        import pandas as pd

        data = [entry.to_export_row() for entry in entries]
        df = pd.DataFrame(data)

        # Reorder columns
        column_order = [name for name, _ in self.COLUMNS if name]
        df = df[[c for c in column_order if c in df.columns]]

        return df


def export_entries(
    entries: list[TimesheetEntry],
    output_path: Path | str,
    template_path: Optional[Path | str] = None,
    append: bool = False,
) -> Path:
    """
    Convenience function to export entries.

    Args:
        entries: List of timesheet entries to export
        output_path: Path for the output file
        template_path: Optional template file to use
        append: If True, append to existing file at output_path

    Returns:
        Path to the exported file
    """
    exporter = ExcelExporter()

    if append and Path(output_path).exists():
        return exporter.append_to_existing(entries, output_path)
    elif template_path:
        return exporter.export_to_template(entries, template_path, output_path)
    else:
        return exporter.export_to_new_file(entries, output_path)


def get_monthly_export_path(
    target_date: Optional[datetime] = None,
    entry_count: Optional[int] = None
) -> Path:
    """
    Get the path for the monthly timesheet file.

    Creates the TIQ Timesheets directory in Documents if it doesn't exist.

    Args:
        target_date: Date to use for the filename (default: today)
        entry_count: Number of entries to include in filename (optional)

    Returns:
        Path to the monthly timesheet file (e.g., Timesheet_January_2026_45entries.xlsx)
    """
    import os

    if target_date is None:
        target_date = datetime.now()

    # Create export directory
    export_dir = Path(os.path.expanduser("~/Documents/TIQ Timesheets"))
    export_dir.mkdir(parents=True, exist_ok=True)

    # Generate filename with month name, year, and entry count
    month_name = target_date.strftime("%B")  # Full month name (e.g., "January")
    year = target_date.strftime("%Y")

    if entry_count is not None:
        filename = f"Timesheet_{month_name}_{year}_{entry_count}entries.xlsx"
    else:
        filename = f"Timesheet_{month_name}_{year}.xlsx"

    return export_dir / filename


def export_to_monthly_file(entries: list[TimesheetEntry]) -> Path:
    """
    Export entries to the monthly timesheet file.

    Creates a new file if it doesn't exist, or appends to existing file.

    Args:
        entries: List of timesheet entries to export

    Returns:
        Path to the exported file
    """
    export_path = get_monthly_export_path()

    if export_path.exists():
        return export_entries(entries, export_path, append=True)
    else:
        return export_entries(entries, export_path)
